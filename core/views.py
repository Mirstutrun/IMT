from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Sum, Q
from .models import UserProfile, Production, Excursion, BarItem, Exhibition, OnlineEvent, Broadcast
import openpyxl
from django.http import HttpResponse

def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        if not username or not password:
            messages.error(request, 'Введите логин и пароль')
            return render(request, 'login.html')
        user = authenticate(request, username=username, password=password)
        if user is None:
            messages.error(request, 'Неправильный логин или пароль')
            return render(request, 'login.html')
        try:
            profile = user.profile
            if not profile.is_active:
                messages.error(request, 'Учётная запись деактивирована')
                return render(request, 'login.html')
        except UserProfile.DoesNotExist:
            pass
        login(request, user)
        return redirect('home')
    return render(request, 'login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


def get_role(request):
    try:
        return request.user.profile.role
    except UserProfile.DoesNotExist:
        return 'viewer'


@login_required
def home_view(request):
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    excursions = Excursion.objects.all()
    exhibitions = Exhibition.objects.all()
    online_events = OnlineEvent.objects.all()
    broadcasts = Broadcast.objects.all()

    if date_from:
        excursions = excursions.filter(date_start__gte=date_from)
        exhibitions = exhibitions.filter(date_start__gte=date_from)
        online_events = online_events.filter(date_start__gte=date_from)
        broadcasts = broadcasts.filter(date_start__gte=date_from)
    if date_to:
        excursions = excursions.filter(date_start__lte=date_to)
        exhibitions = exhibitions.filter(date_start__lte=date_to)
        online_events = online_events.filter(date_start__lte=date_to)
        broadcasts = broadcasts.filter(date_start__lte=date_to)

    stats = {
        'excursions_count': excursions.count(),
        'excursions_participants': excursions.aggregate(s=Sum('participants_total'))['s'] or 0,
        'excursions_pushkin': excursions.aggregate(s=Sum('participants_pushkin'))['s'] or 0,
        'excursions_sum': excursions.aggregate(s=Sum('total_sum'))['s'] or 0,
        'theater_plus_count': excursions.filter(excursion_type='theater_plus').count(),
        'group_show_count': excursions.filter(excursion_type='group_show').count(),
        'group_excursion_count': excursions.filter(excursion_type='group_excursion').count(),
        'exhibitions_count': exhibitions.count(),
        'online_count': online_events.count(),
        'broadcasts_count': broadcasts.count(),
    }
    return render(request, 'home.html', {'stats': stats, 'date_from': date_from, 'date_to': date_to})


@login_required
def productions_view(request):
    role = get_role(request)
    productions = Production.objects.all()
    search = request.GET.get('search', '')
    if search:
        productions = productions.filter(Q(title__icontains=search) | Q(genre__icontains=search))

    if request.method == 'POST' and role in ('admin', 'excursion'):
        action = request.POST.get('action')
        if action == 'create':
            title = request.POST.get('title', '').strip()
            genre = request.POST.get('genre', '').strip()
            if title and genre:
                if Production.objects.filter(title__iexact=title).exists():
                    messages.error(request, 'Постановка с таким названием уже существует')
                else:
                    Production.objects.create(title=title, genre=genre)
                    messages.success(request, 'Постановка добавлена')
            else:
                messages.error(request, 'Заполните все поля')
        elif action == 'edit':
            pk = request.POST.get('pk')
            p = get_object_or_404(Production, pk=pk)
            p.title = request.POST.get('title', p.title).strip()
            p.genre = request.POST.get('genre', p.genre).strip()
            p.save()
            messages.success(request, 'Постановка обновлена')
        elif action == 'delete':
            pk = request.POST.get('pk')
            Production.objects.filter(pk=pk).delete()
            messages.success(request, 'Постановка удалена')
        return redirect('productions')

    return render(request, 'productions.html', {'productions': productions, 'role': role, 'search': search})


@login_required
def excursions_view(request):
    role = get_role(request)
    excursions = Excursion.objects.select_related('production').all()
    search = request.GET.get('search', '')
    type_filter = request.GET.get('type_filter', '')
    status_filter = request.GET.get('status_filter', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if search:
        excursions = excursions.filter(Q(organization__icontains=search) | Q(contact_person__icontains=search))
    if type_filter:
        excursions = excursions.filter(excursion_type=type_filter)
    if status_filter:
        excursions = excursions.filter(status=status_filter)

    report_data = None
    if date_from and date_to:
        report_qs = Excursion.objects.filter(date_start__gte=date_from, date_start__lte=date_to)
        report_data = {
            'total': report_qs.count(),
            'participants': report_qs.aggregate(s=Sum('participants_total'))['s'] or 0,
            'pushkin': report_qs.aggregate(s=Sum('participants_pushkin'))['s'] or 0,
            'sum': report_qs.aggregate(s=Sum('total_sum'))['s'] or 0,
            'theater_plus': report_qs.filter(excursion_type='theater_plus').count(),
            'group_show': report_qs.filter(excursion_type='group_show').count(),
            'group_excursion': report_qs.filter(excursion_type='group_excursion').count(),
        }

    if request.method == 'POST' and role in ('admin', 'excursion'):
        action = request.POST.get('action')
        if action == 'create':
            exc_type = request.POST.get('excursion_type')
            prod_id = request.POST.get('production') or None
            org = request.POST.get('organization', '').strip()
            contact = request.POST.get('contact_person', '').strip()
            date_s = request.POST.get('date_start')
            date_e = request.POST.get('date_end') or None
            participants = int(request.POST.get('participants_total') or 0)
            pushkin = int(request.POST.get('participants_pushkin') or 0)
            total = request.POST.get('total_sum') or 0
            has_bar = request.POST.get('has_bar') == 'on'
            status = request.POST.get('status', 'new')
            if not org or not contact or not date_s:
                messages.error(request, 'Заполните обязательные поля')
            else:
                try:
                    exc = Excursion.objects.create(
                        organization=org, contact_person=contact,
                        excursion_type=exc_type, production_id=prod_id,
                        date_start=date_s, date_end=date_e,
                        participants_total=participants, participants_pushkin=pushkin,
                        total_sum=total, has_bar=has_bar, status=status
                    )
                    if has_bar:
                        bar_names = request.POST.getlist('bar_name')
                        bar_prices = request.POST.getlist('bar_price')
                        for n, p in zip(bar_names, bar_prices):
                            if n.strip():
                                BarItem.objects.create(excursion=exc, name=n.strip(), price=p or 0)
                    messages.success(request, 'Заявка создана')
                except Exception as e:
                    messages.error(request, f'Ошибка при сохранении: {e}')
        elif action == 'edit':
            pk = request.POST.get('pk')
            exc = get_object_or_404(Excursion, pk=pk)
            exc.organization = request.POST.get('organization', exc.organization).strip()
            exc.contact_person = request.POST.get('contact_person', exc.contact_person).strip()
            exc.excursion_type = request.POST.get('excursion_type', exc.excursion_type)
            prod_id = request.POST.get('production') or None
            exc.production_id = prod_id
            exc.date_start = request.POST.get('date_start', exc.date_start)
            exc.date_end = request.POST.get('date_end') or None
            exc.participants_total = int(request.POST.get('participants_total') or exc.participants_total)
            exc.participants_pushkin = int(request.POST.get('participants_pushkin') or exc.participants_pushkin)
            exc.total_sum = request.POST.get('total_sum') or exc.total_sum
            exc.has_bar = request.POST.get('has_bar') == 'on'
            exc.status = request.POST.get('status', exc.status)
            exc.save()
            messages.success(request, 'Заявка обновлена')
        elif action == 'delete':
            pk = request.POST.get('pk')
            Excursion.objects.filter(pk=pk).delete()
            messages.success(request, 'Заявка удалена')
        return redirect('excursions')

    productions = Production.objects.all()
    return render(request, 'excursions.html', {
        'excursions': excursions, 'role': role, 'productions': productions,
        'search': search, 'type_filter': type_filter, 'status_filter': status_filter,
        'date_from': date_from, 'date_to': date_to, 'report_data': report_data,
    })


@login_required
def exhibitions_view(request):
    role = get_role(request)
    exhibitions = Exhibition.objects.all()
    search = request.GET.get('search', '')
    category_filter = request.GET.get('category_filter', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if search:
        exhibitions = exhibitions.filter(Q(title__icontains=search) | Q(location__icontains=search) | Q(contact_person__icontains=search))
    if category_filter:
        exhibitions = exhibitions.filter(category=category_filter)

    report_data = None
    if date_from and date_to:
        report_qs = Exhibition.objects.filter(date_start__gte=date_from, date_start__lte=date_to)
        report_data = {
            'total': report_qs.count(),
            'works': report_qs.aggregate(s=Sum('works_count'))['s'] or 0,
        }

    if request.method == 'POST' and role in ('admin', 'excursion'):
        action = request.POST.get('action')
        if action == 'create':
            title = request.POST.get('title', '').strip()
            category = request.POST.get('category')
            location = request.POST.get('location', '').strip()
            date_s = request.POST.get('date_start')
            date_e = request.POST.get('date_end')
            works = request.POST.get('works_count') or 0
            contact = request.POST.get('contact_person', '').strip()
            status = request.POST.get('status', 'planned')
            if not title or not date_s or not date_e:
                messages.error(request, 'Заполните обязательные поля')
            else:
                try:
                    Exhibition.objects.create(
                        title=title, category=category, location=location,
                        date_start=date_s, date_end=date_e, works_count=works,
                        contact_person=contact, status=status
                    )
                    messages.success(request, 'Выставка добавлена')
                except Exception as e:
                    messages.error(request, f'Ошибка при сохранении: {e}')
        elif action == 'edit':
            pk = request.POST.get('pk')
            ex = get_object_or_404(Exhibition, pk=pk)
            ex.title = request.POST.get('title', ex.title).strip()
            ex.category = request.POST.get('category', ex.category)
            ex.location = request.POST.get('location', ex.location).strip()
            ex.date_start = request.POST.get('date_start', ex.date_start)
            ex.date_end = request.POST.get('date_end', ex.date_end)
            ex.works_count = request.POST.get('works_count') or ex.works_count
            ex.contact_person = request.POST.get('contact_person', ex.contact_person).strip()
            ex.status = request.POST.get('status', ex.status)
            ex.save()
            messages.success(request, 'Выставка обновлена')
        elif action == 'delete':
            pk = request.POST.get('pk')
            Exhibition.objects.filter(pk=pk).delete()
            messages.success(request, 'Выставка удалена')
        return redirect('exhibitions')

    return render(request, 'exhibitions.html', {
        'exhibitions': exhibitions, 'role': role,
        'search': search, 'category_filter': category_filter,
        'date_from': date_from, 'date_to': date_to, 'report_data': report_data,
    })


@login_required
def online_events_view(request):
    role = get_role(request)
    events = OnlineEvent.objects.all()
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status_filter', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if search:
        events = events.filter(title__icontains=search)
    if status_filter:
        events = events.filter(status=status_filter)

    report_data = None
    if date_from and date_to:
        report_qs = OnlineEvent.objects.filter(date_start__gte=date_from, date_start__lte=date_to)
        report_data = {
            'total': report_qs.count(),
        }

    if request.method == 'POST' and role in ('admin', 'excursion'):
        action = request.POST.get('action')
        if action == 'create':
            title = request.POST.get('title', '').strip()
            date_s = request.POST.get('date_start')
            date_e = request.POST.get('date_end') or None
            status = request.POST.get('status', 'planned')
            if not title or not date_s:
                messages.error(request, 'Заполните обязательные поля')
            else:
                try:
                    OnlineEvent.objects.create(title=title, event_type='online', date_start=date_s, date_end=date_e, status=status)
                    messages.success(request, 'Активность добавлена')
                except Exception as e:
                    messages.error(request, f'Ошибка при сохранении: {e}')
        elif action == 'edit':
            pk = request.POST.get('pk')
            ev = get_object_or_404(OnlineEvent, pk=pk)
            ev.title = request.POST.get('title', ev.title).strip()
            ev.date_start = request.POST.get('date_start', ev.date_start)
            ev.date_end = request.POST.get('date_end') or None
            ev.status = request.POST.get('status', ev.status)
            ev.save()
            messages.success(request, 'Активность обновлена')
        elif action == 'delete':
            pk = request.POST.get('pk')
            OnlineEvent.objects.filter(pk=pk).delete()
            messages.success(request, 'Активность удалена')
        return redirect('online_events')

    return render(request, 'online_events.html', {
        'events': events, 'role': role,
        'search': search, 'status_filter': status_filter,
        'date_from': date_from, 'date_to': date_to, 'report_data': report_data,
    })


@login_required
def broadcasts_view(request):
    role = get_role(request)
    broadcasts = Broadcast.objects.all()
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status_filter', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if search:
        broadcasts = broadcasts.filter(title__icontains=search)
    if status_filter:
        broadcasts = broadcasts.filter(status=status_filter)

    report_data = None
    if date_from and date_to:
        report_qs = Broadcast.objects.filter(date_start__gte=date_from, date_start__lte=date_to)
        report_data = {
            'total': report_qs.count(),
        }

    if request.method == 'POST' and role in ('admin', 'excursion'):
        action = request.POST.get('action')
        if action == 'create':
            title = request.POST.get('title', '').strip()
            date_s = request.POST.get('date_start')
            date_e = request.POST.get('date_end') or None
            status = request.POST.get('status', 'planned')
            if not title or not date_s:
                messages.error(request, 'Заполните обязательные поля')
            else:
                try:
                    Broadcast.objects.create(title=title, date_start=date_s, date_end=date_e, status=status)
                    messages.success(request, 'Трансляция добавлена')
                except Exception as e:
                    messages.error(request, f'Ошибка: {e}')
        elif action == 'edit':
            pk = request.POST.get('pk')
            br = get_object_or_404(Broadcast, pk=pk)
            br.title = request.POST.get('title', br.title).strip()
            br.date_start = request.POST.get('date_start', br.date_start)
            br.date_end = request.POST.get('date_end') or None
            br.status = request.POST.get('status', br.status)
            br.save()
            messages.success(request, 'Трансляция обновлена')
        elif action == 'delete':
            pk = request.POST.get('pk')
            Broadcast.objects.filter(pk=pk).delete()
            messages.success(request, 'Трансляция удалена')
        return redirect('broadcasts')

    return render(request, 'broadcasts.html', {
        'broadcasts': broadcasts, 'role': role,
        'search': search, 'status_filter': status_filter,
        'date_from': date_from, 'date_to': date_to, 'report_data': report_data,
    })


@login_required
def bar_items_view(request):
    role = get_role(request)
    items = BarItem.objects.select_related('excursion').all()
    search = request.GET.get('search', '')
    if search:
        items = items.filter(Q(name__icontains=search) | Q(excursion__organization__icontains=search))

    if request.method == 'POST' and role in ('admin', 'excursion'):
        action = request.POST.get('action')
        if action == 'create':
            exc_id = request.POST.get('excursion')
            name = request.POST.get('name', '').strip()
            price = request.POST.get('price') or 0
            if not name or not exc_id:
                messages.error(request, 'Заполните все поля')
            else:
                BarItem.objects.create(excursion_id=exc_id, name=name, price=price)
                messages.success(request, 'Позиция добавлена')
        elif action == 'edit':
            pk = request.POST.get('pk')
            item = get_object_or_404(BarItem, pk=pk)
            item.excursion_id = request.POST.get('excursion', item.excursion_id)
            item.name = request.POST.get('name', item.name).strip()
            item.price = request.POST.get('price') or item.price
            item.save()
            messages.success(request, 'Позиция обновлена')
        elif action == 'delete':
            pk = request.POST.get('pk')
            BarItem.objects.filter(pk=pk).delete()
            messages.success(request, 'Позиция удалена')
        return redirect('bar_items')

    excursions = Excursion.objects.filter(has_bar=True)
    return render(request, 'bar_items.html', {'items': items, 'role': role, 'excursions': excursions, 'search': search})

@login_required
def export_excel(request):
    date_from = request.GET.get('date_from') or None
    date_to = request.GET.get('date_to') or None

    excursions = Excursion.objects.select_related('production').all()
    exhibitions = Exhibition.objects.all()
    online_events = OnlineEvent.objects.all()
    broadcasts = Broadcast.objects.all()

    if date_from:
        excursions = excursions.filter(date_start__gte=date_from)
        exhibitions = exhibitions.filter(date_start__gte=date_from)
        online_events = online_events.filter(date_start__gte=date_from)
        broadcasts = broadcasts.filter(date_start__gte=date_from)
    if date_to:
        excursions = excursions.filter(date_start__lte=date_to)
        exhibitions = exhibitions.filter(date_start__lte=date_to)
        online_events = online_events.filter(date_start__lte=date_to)
        broadcasts = broadcasts.filter(date_start__lte=date_to)

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = 'Экскурсии'
    ws1.append(['Организация', 'Контакт', 'Тип', 'Постановка', 'Дата начала', 'Дата конца', 'Участников', 'По Пушк. карте', 'Сумма', 'Бар', 'Статус'])
    for e in excursions:
        ws1.append([
            e.organization,
            e.contact_person,
            e.get_excursion_type_display(),
            e.production.title if e.production else '',
            str(e.date_start),
            str(e.date_end) if e.date_end else '',
            e.participants_total,
            e.participants_pushkin,
            float(e.total_sum),
            'Да' if e.has_bar else 'Нет',
            e.get_status_display(),
        ])

    ws2 = wb.create_sheet('Выставки')
    ws2.append(['Название', 'Категория', 'Место', 'Дата начала', 'Дата конца', 'Кол-во работ', 'Контакт', 'Статус'])
    for ex in exhibitions:
        ws2.append([
            ex.title,
            ex.get_category_display(),
            ex.location,
            str(ex.date_start),
            str(ex.date_end),
            ex.works_count,
            ex.contact_person,
            ex.get_status_display(),
        ])

    ws3 = wb.create_sheet('Трансляции')
    ws3.append(['Название', 'Дата начала', 'Дата конца', 'Статус'])
    for br in broadcasts:
        ws3.append([
            br.title,
            str(br.date_start),
            str(br.date_end) if br.date_end else '',
            br.get_status_display(),
        ])

    ws4 = wb.create_sheet('Онлайн-активности')
    ws4.append(['Название', 'Дата начала', 'Дата конца', 'Статус'])
    for ev in online_events:
        ws4.append([
            ev.title,
            str(ev.date_start),
            str(ev.date_end) if ev.date_end else '',
            ev.get_status_display(),
        ])

    ws5 = wb.create_sheet('Сводный отчёт')
    ws5.append(['Показатель', 'Значение'])
    ws5.append(['Период с', date_from or 'всё время'])
    ws5.append(['Период по', date_to or 'всё время'])
    ws5.append([])
    ws5.append(['Экскурсий всего', excursions.count()])
    ws5.append(['— Театр+', excursions.filter(excursion_type='theater_plus').count()])
    ws5.append(['— Групповой спектакль', excursions.filter(excursion_type='group_show').count()])
    ws5.append(['— Групповая экскурсия', excursions.filter(excursion_type='group_excursion').count()])
    ws5.append(['Участников всего', excursions.aggregate(s=Sum('participants_total'))['s'] or 0])
    ws5.append(['По Пушкинской карте', excursions.aggregate(s=Sum('participants_pushkin'))['s'] or 0])
    ws5.append(['Сумма заявок (руб.)', float(excursions.aggregate(s=Sum('total_sum'))['s'] or 0)])
    ws5.append([])
    ws5.append(['Выставок', exhibitions.count()])
    ws5.append(['Трансляций', broadcasts.count()])
    ws5.append(['Онлайн-активностей', online_events.count()])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="report.xlsx"'
    wb.save(response)
    return response

@login_required
def users_view(request):
    role = get_role(request)
    if role != 'admin':
        messages.error(request, 'Доступ запрещён')
        return redirect('home')

    users = UserProfile.objects.select_related('user').all()
    search = request.GET.get('search', '')
    if search:
        users = users.filter(Q(full_name__icontains=search) | Q(user__username__icontains=search))

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'create':
            username = request.POST.get('username', '').strip()
            full_name = request.POST.get('full_name', '').strip()
            password = request.POST.get('password', '').strip()
            user_role = request.POST.get('role', 'viewer')
            if not username or not full_name or not password:
                messages.error(request, 'Заполните все поля')
            elif User.objects.filter(username=username).exists():
                messages.error(request, 'Логин уже занят')
            else:
                u = User.objects.create_user(username=username, password=password)
                UserProfile.objects.create(user=u, full_name=full_name, role=user_role)
                messages.success(request, 'Пользователь создан')
        elif action == 'edit':
            pk = request.POST.get('pk')
            profile = get_object_or_404(UserProfile, pk=pk)
            new_username = request.POST.get('username', '').strip()
            if new_username and new_username != profile.user.username:
                if User.objects.filter(username=new_username).exclude(pk=profile.user.pk).exists():
                    messages.error(request, 'Логин уже занят')
                    return redirect('users')
                profile.user.username = new_username
                profile.user.save()
            new_password = request.POST.get('password', '').strip()
            if new_password:
                profile.user.set_password(new_password)
                profile.user.save()
            profile.full_name = request.POST.get('full_name', profile.full_name).strip()
            profile.role = request.POST.get('role', profile.role)
            profile.save()
            messages.success(request, 'Пользователь обновлён')
        elif action == 'delete':
            pk = request.POST.get('pk')
            profile = get_object_or_404(UserProfile, pk=pk)
            profile.is_active = False
            profile.user.is_active = False
            profile.user.save()
            profile.save()
            messages.success(request, 'Пользователь деактивирован')
        return redirect('users')

    return render(request, 'users.html', {'users': users, 'role': role, 'search': search})